"""
Excel Connector for parsing Excel (.xlsx) files.
"""

from pathlib import Path
from typing import Iterator, Optional
from openpyxl import load_workbook
from .base import BaseConnector, DataRecord
from ..core.exception import ConnectorError


class ExcelConnector(BaseConnector):
    """
    Connector for Excel files.
    
    Features:
    - Multiple sheet support
    - Auto-detect header row
    - Type preservation from Excel
    
    Example:
        connector = ExcelConnector()
        for record in connector.parse("products.xlsx"):
            product_id = record.get("product_id")
    """
    
    def __init__(
        self,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
        skip_rows: int = 0
    ):
        """
        Initialize Excel connector.
        
        Args:
            sheet_name: Sheet to read (first sheet if None)
            header_row: Row number containing headers (1-indexed)
            skip_rows: Number of data rows to skip after header
        """
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.skip_rows = skip_rows
    
    def parse(self, source: Path | str) -> Iterator[DataRecord]:
        """
        Parse Excel file and yield records.
        
        Args:
            source: Path to Excel file
            
        Yields:
            DataRecord objects (one per row)
            
        Raises:
            ConnectorError: If Excel parsing fails
        """
        try:
            source_path = Path(source) if isinstance(source, str) else source
            
            wb = load_workbook(source_path, read_only=True, data_only=True)
            
            # Get sheet
            if self.sheet_name:
                if self.sheet_name not in wb.sheetnames:
                    raise ConnectorError(
                        f"Sheet '{self.sheet_name}' not found. "
                        f"Available: {wb.sheetnames}"
                    )
                sheet = wb[self.sheet_name]
            else:
                sheet = wb.active
            
            # Get headers from specified row
            headers = []
            for row in sheet.iter_rows(
                min_row=self.header_row, 
                max_row=self.header_row,
                values_only=True
            ):
                headers = [str(h).strip() if h else f"column_{i}" 
                          for i, h in enumerate(row)]
                break
            
            if not headers:
                raise ConnectorError("No headers found in Excel file")
            
            # Read data rows
            start_row = self.header_row + 1 + self.skip_rows
            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                # Skip empty rows
                if all(cell is None for cell in row):
                    continue
                    
                # Create record with headers as keys
                data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        # Preserve type from Excel, convert to string if needed
                        data[headers[i]] = value
                
                yield DataRecord(data=data)
            
            wb.close()
            
        except FileNotFoundError:
            raise ConnectorError(f"Excel file not found: {source}")
        except Exception as e:
            if isinstance(e, ConnectorError):
                raise
            raise ConnectorError(f"Failed to parse Excel: {e}")
    
    def get_schema(self) -> dict:
        """Describe Excel connector output"""
        return {
            "connector": "ExcelConnector",
            "description": "Excel File Parser",
            "sheet_name": self.sheet_name or "(active sheet)",
            "header_row": self.header_row
        }
